# fyp-backend/projects/views.py
from rest_framework import viewsets, status, mixins, filters, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import JSONParser, MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django.utils import timezone
from django.core.exceptions import PermissionDenied
from accounts.models import CustomUser 
from django.db import models
from .models import ProjectGroup, Faculty
from .serializers import ProjectGroupSerializer
from .models import MeetingMinute, AttendanceLog


from .models import ProjectGroup, GroupMember, Faculty, FYDPProposal, ChangeRequest, MeetingMinute, AttendanceLog, Announcement
from .serializers import (
    FacultyListSerializer,
    GroupMemberSerializer,
    ProjectGroupSerializer,
    GroupCreateSerializer,
    FYDPProposalSerializer,
    FYDPProposalUploadSerializer,  
    FYDPProposalReviewSerializer,     
    ChangeRequestSerializer,
    AdminProjectGroupSerializer,
    AdminApprovalSerializer,
    MeetingMinuteSerializer,
    MeetingMinuteCreateUpdateSerializer,
    AttendanceLogSerializer, 
    AnnouncementSerializer, 
    AnnouncementCreateUpdateSerializer, 
)
from .permissions import IsStudent, IsGroupMemberOrReadOnly, IsAdminUser


# =============================================================================
# 1. FACULTY LIST - Dropdown ke liye (Read Only)
# =============================================================================
class FacultyViewSet(mixins.ListModelMixin, mixins.RetrieveModelMixin, viewsets.GenericViewSet):
    """Show ALL active faculty. No filtering/blocking."""
    queryset = Faculty.objects.filter(is_active=True)
    serializer_class = FacultyListSerializer
    permission_classes = [IsAuthenticated]


# =============================================================================
# 2. ELIGIBILITY CHECK - Real-time warning preview
# =============================================================================
class EligibilityCheckView(viewsets.ViewSet):
    """
    POST /api/projects/check-eligibility/
    Body: { "cgpa": 1.8, "earned_credit_hours": 95, "prerequisites_completed": true }
    Returns warnings but does NOT save anything.
    """
    permission_classes = [IsAuthenticated]

    def create(self, request, *args, **kwargs):
        members_data = request.data.pop('members', [])
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        self.perform_create(serializer)
        group = serializer.instance
        
        member_errors = []
        for idx, member_data in enumerate(members_data):
            user_odoo_id = member_data.get('odoo_id')
            user = CustomUser.objects.filter(student_id=user_odoo_id).first()
            
            if not user:
                member_errors.append({
                    "index": idx,
                    "error": f"User with ID {user_odoo_id} not found in database."
                })
                continue
            
            try:
                GroupMember.objects.create(
                    group=group,
                    student=user,
                    role=member_data.get('role', 'member'),
                    full_name=member_data.get('full_name'),
                    odoo_id=member_data.get('odoo_id'),
                    cgpa=member_data.get('cgpa'),
                    earned_credit_hours=member_data.get('earned_credit_hours'),
                    prerequisites_completed=member_data.get('prerequisites_completed', True),
                    has_special_permission=member_data.get('has_special_permission', False)
                )
            except Exception as e:
                member_errors.append({"index": idx, "error": str(e)})

        if member_errors:
            group.delete()
            return Response({"member_errors": member_errors}, status=status.HTTP_400_BAD_REQUEST)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


# =============================================================================
# 3. PROJECT GROUPS - Creation & Management
# =============================================================================
class ProjectGroupViewSet(viewsets.ModelViewSet):
    queryset = ProjectGroup.objects.all()
    serializer_class = ProjectGroupSerializer
    parser_classes = [JSONParser, MultiPartParser, FormParser]

    def create(self, request, *args, **kwargs):
        # ✅ Save members data first
        members_data = request.data.pop('members', [])
        
        # ✅ NO AUTO-DELETE: Rejected groups will stay in database for audit trail
        # Admin can see all rejected groups in history
        
        # 2️⃣ Continue with normal group creation
        request.data['status'] = 'pending_approval'
        request.data.pop('group_number', None)
        
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        group = serializer.instance
        
        print(f"✅ New group created: {group.id}, Status: {group.status}")
        
        # 3️⃣ Create members
        member_errors = []
        for idx, member_data in enumerate(members_data):
            user_input_id = member_data.get('odoo_id')
            user = CustomUser.objects.filter(student_id=user_input_id).first()
            
            if not user:
                member_errors.append({
                    "index": idx,
                    "error": f"User '{user_input_id}' not found."
                })
                continue
            
            GroupMember.objects.create(
                group=group,
                student=user,
                role=member_data.get('role', 'member'),
                full_name=member_data.get('full_name'),
                odoo_id=member_data.get('odoo_id'),
                cgpa=member_data.get('cgpa'),
                earned_credit_hours=member_data.get('earned_credit_hours'),
                prerequisites_completed=member_data.get('prerequisites_completed', True),
                has_special_permission=member_data.get('has_special_permission', False)
            )
            
            print(f"✅ Member {idx+1} added: {user.email}")
        
        if member_errors:
            group.delete()
            return Response({"member_errors": member_errors}, status=status.HTTP_400_BAD_REQUEST)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
    
    @action(detail=False, methods=['get'])
    def my_group(self, request):
        """Get current user's group (EXCLUDE rejected groups)"""
        user = request.user
        
        # ✅ Only return non-rejected groups
        member = GroupMember.objects.filter(
            student=user,
            group__status__in=[
                'pending_approval',
                'idea_pitch',
                'approved',
                'proposal_pending',
                'proposal_approved',
                'in_progress',
                'completed'
            ]
        ).select_related('group').first()
        
        if member:
            serializer = self.get_serializer(member.group)
            return Response(serializer.data)
        
        # No active group found - student can create new one
        return Response(None, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['get'])
    def pending_approval(self, request):
        """
        GET /api/projects/groups/pending_approval/
        For students to see their pending groups
        """
        user = request.user
        pending = ProjectGroup.objects.filter(
            members__student=user,
            status='pending_approval'
        ).prefetch_related('members__student')
        
        serializer = self.get_serializer(pending, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'], url_path='my-group-with-history')
    def my_group_with_history(self, request):
        """
        Get user's latest group (including rejected ones for display)
        This helps show rejection status to students
        """
        user = request.user
        
        # Get ALL groups for this user, ordered by creation date
        member = GroupMember.objects.filter(
            student=user
        ).select_related('group').order_by('-group__created_at').first()
        
        if member:
            serializer = self.get_serializer(member.group)
            return Response(serializer.data)
        
        return Response(None, status=status.HTTP_404_NOT_FOUND)


# =============================================================================
# 4. FYDP PROPOSALS - Digital Form 1 & File Upload
# =============================================================================
class FYDPProposalViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing FYDP Proposals.
    Handles student uploads, supervisor reviews, and admin final approvals.
    """
    serializer_class = FYDPProposalSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        """
        Filter proposals based on user role:
        - Students: Only their own group's proposal
        - Supervisors: Proposals of their assigned groups
        - Admins: All proposals
        """
        user = self.request.user
        queryset = FYDPProposal.objects.all()

        if user.user_type == 'student':
            return queryset.filter(group__members__student=user)
        elif user.user_type == 'supervisor':
            try:
                faculty = user.faculty_profile
                return queryset.filter(
                    models.Q(group__supervisor=faculty) | models.Q(group__co_supervisor=faculty)
                ).distinct()
            except Faculty.DoesNotExist:
                return FYDPProposal.objects.none()
        elif user.user_type == 'admin':
            return queryset
        
        return FYDPProposal.objects.none()

    def get_serializer_class(self):
        if self.action == 'upload':
            return FYDPProposalUploadSerializer
        if self.action in ['supervisor_review', 'admin_review']:
            return FYDPProposalReviewSerializer
        return FYDPProposalSerializer

    @action(detail=False, methods=['get'], url_path='pending-supervisor')
    def pending_supervisor(self, request):
        """
        GET /api/projects/proposals/pending-supervisor/
        For supervisors to see proposals pending their review.
        """
        if request.user.user_type != 'supervisor':
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)

        try:
            faculty = request.user.faculty_profile
        except Faculty.DoesNotExist:
            return Response({'error': 'Faculty profile not found'}, status=status.HTTP_404_NOT_FOUND)

        pending_proposals = self.get_queryset().filter(
            status__in=['submitted', 'revision_needed']
        )
        
        serializer = self.get_serializer(pending_proposals, many=True)
        return Response({'count': pending_proposals.count(), 'results': serializer.data})

    @action(detail=False, methods=['get'], url_path='pending-admin')
    def pending_admin(self, request):
        """
        GET /api/projects/proposals/pending-admin/
        For admins to see proposals pending final approval.
        """
        if request.user.user_type != 'admin':
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)

        pending_proposals = FYDPProposal.objects.filter(status='approved_by_supervisor')
        serializer = self.get_serializer(pending_proposals, many=True)
        return Response({'count': pending_proposals.count(), 'results': serializer.data})

    @action(detail=True, methods=['post'], url_path='upload')
    def upload(self, request, pk=None):
        """
        POST /api/projects/proposals/{id}/upload/
        Student uploads the proposal file (PDF/DOCX).
        """
        if request.user.user_type != 'student':
            return Response({'error': 'Only students can upload proposals'}, status=status.HTTP_403_FORBIDDEN)

        proposal = self.get_object()
        
        # Check if upload is allowed
        can_upload, reason = proposal.can_upload()
        if not can_upload:
            return Response({'error': reason}, status=status.HTTP_400_BAD_REQUEST)

        # Validate file
        serializer = FYDPProposalUploadSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        # Save file and update proposal
        proposal.proposal_file = serializer.validated_data['proposal_file']
        proposal.status = 'submitted'
        proposal.submitted_at = timezone.now()
        proposal.supervisor_remarks = '' # Clear old remarks if re-uploading after revision
        proposal.admin_remarks = ''
        
        # Increment submission count
        success, msg = proposal.increment_submission_count()
        if not success:
            return Response({'error': msg}, status=status.HTTP_400_BAD_REQUEST)
            
        proposal.save()
        
        return Response({
            'message': 'Proposal uploaded successfully',
            'data': FYDPProposalSerializer(proposal).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='supervisor-review')
    def supervisor_review(self, request, pk=None):
        """
        POST /api/projects/proposals/{id}/supervisor-review/
        Supervisor approves or requests revision.
        """
        if request.user.user_type != 'supervisor':
            return Response({'error': 'Only supervisors can review proposals'}, status=status.HTTP_403_FORBIDDEN)

        proposal = self.get_object()
        
        # Verify supervisor is assigned to this group
        try:
            faculty = request.user.faculty_profile
            if proposal.group.supervisor != faculty and proposal.group.co_supervisor != faculty:
                return Response({'error': 'You are not assigned to this group'}, status=status.HTTP_403_FORBIDDEN)
        except Faculty.DoesNotExist:
            return Response({'error': 'Faculty profile not found'}, status=status.HTTP_404_NOT_FOUND)

        # Validate action
        serializer = FYDPProposalReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        action = serializer.validated_data['action']
        remarks = serializer.validated_data.get('remarks', '')
        
        success, msg = proposal.supervisor_review(faculty, action, remarks)
        if not success:
            return Response({'error': msg}, status=status.HTTP_400_BAD_REQUEST)
            
        return Response({
            'message': msg,
            'data': FYDPProposalSerializer(proposal).data
        }, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='admin-review')
    def admin_review(self, request, pk=None):
        """
        POST /api/projects/proposals/{id}/admin-review/
        Admin gives final approval or rejection.
        """
        if request.user.user_type != 'admin':
            return Response({'error': 'Only admins can give final approval'}, status=status.HTTP_403_FORBIDDEN)

        proposal = self.get_object()
        
        # Validate action
        serializer = FYDPProposalReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        action = serializer.validated_data['action']
        remarks = serializer.validated_data.get('remarks', '')
        
        success, msg = proposal.admin_review(request.user, action, remarks)
        if not success:
            return Response({'error': msg}, status=status.HTTP_400_BAD_REQUEST)
            
        return Response({
            'message': msg,
            'data': FYDPProposalSerializer(proposal).data
        }, status=status.HTTP_200_OK)

# =============================================================================
# 5. CHANGE REQUESTS - Form 3 & 4
# =============================================================================
class ChangeRequestViewSet(viewsets.ModelViewSet):
    serializer_class = ChangeRequestSerializer
    permission_classes = [IsAuthenticated, IsStudent]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return ChangeRequest.objects.filter(requested_by=self.request.user)

    def perform_create(self, serializer):
        serializer.save(requested_by=self.request.user)


# =============================================================================
# ✅ ADMIN APPROVAL VIEWSET - FIXED & COMPLETE
# =============================================================================

class AdminGroupApprovalViewSet(viewsets.ViewSet):
    """
    Admin ke liye group approval system.
    Endpoints:
    - GET /api/projects/admin/approval/pending/
    - GET /api/projects/admin/approval/all/
    - POST /api/projects/admin/approval/{id}/approve/
    - POST /api/projects/admin/approval/{id}/reject/
    - GET /api/projects/admin/approval/{id}/details/
    """
    permission_classes = [IsAuthenticated]
    
    def check_admin_permission(self, request):
        """Check if user is admin"""
        if not hasattr(request.user, 'user_type') or request.user.user_type != 'admin':
            raise PermissionDenied("Only admin users can access this endpoint")
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        """GET /api/projects/admin/approval/pending/"""
        self.check_admin_permission(request)
        
        pending_groups = ProjectGroup.objects.filter(
            status='pending_approval'
        ).prefetch_related(
            'members__student',
            'supervisor',
            'co_supervisor'
        ).order_by('-created_at')
        
        serializer = AdminProjectGroupSerializer(pending_groups, many=True)
        return Response({
            'count': pending_groups.count(),
            'results': serializer.data
        })
    
    @action(detail=False, methods=['get'], url_path='all')
    def all_groups(self, request):
        """GET /api/projects/admin/approval/all/"""
        self.check_admin_permission(request)
        
        groups = ProjectGroup.objects.all().prefetch_related(
            'members__student',
            'supervisor',
            'co_supervisor'
        ).order_by('-created_at')
        
        # Optional filtering
        status_filter = request.query_params.get('status', None)
        if status_filter:
            groups = groups.filter(status=status_filter)
        
        semester_filter = request.query_params.get('semester', None)
        if semester_filter:
            groups = groups.filter(semester=semester_filter)
        
        paginator = PageNumberPagination()
        paginator.page_size = 20
        result_page = paginator.paginate_queryset(groups, request)
        
        serializer = AdminProjectGroupSerializer(result_page, many=True)
        return paginator.get_paginated_response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """POST /api/projects/admin/approval/{id}/approve/"""
        self.check_admin_permission(request)
        
        try:
            group = ProjectGroup.objects.select_related(
                'supervisor', 'co_supervisor'
            ).prefetch_related('members__student').get(pk=pk)
        except ProjectGroup.DoesNotExist:
            return Response(
                {'error': 'Group not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        if group.status != 'pending_approval':
            return Response(
                {'error': f'Group is not pending approval. Current status: {group.status}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check for optional group number override
        override_number = request.data.get('group_number_override', None)
        if override_number:
            if ProjectGroup.objects.filter(group_number=override_number).exclude(pk=pk).exists():
                return Response(
                    {'error': 'Group number already exists'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            group.group_number = override_number
        
        # ✅ Approve logic inline (no need for model method)
        group.status = 'idea_pitch'  # Next status after approval
        group.approved_by = request.user
        group.approved_at = timezone.now()
        group.rejection_reason = None
        
        # Auto-generate group_number if not provided
        if not group.group_number:
            semester = group.semester or '2026'
            year = semester.split()[0] if semester else '2026'
            last_group = ProjectGroup.objects.filter(
                semester=semester,
                group_number__isnull=False,
                group_number__startswith=f'GRP-{year}'
            ).order_by('-group_number').first()
            
            if last_group and last_group.group_number:
                try:
                    last_num = int(last_group.group_number.split('-')[-1])
                    group.group_number = f"GRP-{year}-{last_num + 1:03d}"
                except:
                    group.group_number = f"GRP-{year}-001"
            else:
                group.group_number = f"GRP-{year}-001"
        
        group.save()
        
        serializer = AdminProjectGroupSerializer(group)
        return Response({
            'message': f'Group {group.group_number} approved successfully!',
            'data': serializer.data
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """POST /api/projects/admin/approval/{id}/reject/"""
        self.check_admin_permission(request)
        
        reason = request.data.get('reason', '').strip()
        if not reason:
            return Response(
                {'error': 'Rejection reason is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            group = ProjectGroup.objects.get(pk=pk)
        except ProjectGroup.DoesNotExist:
            return Response(
                {'error': 'Group not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        if group.status not in ['pending_approval']:
            return Response(
                {'error': f'Group cannot be rejected. Current status: {group.status}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # ✅ Reject logic inline
        group.status = 'rejected'
        group.rejection_reason = reason
        group.approved_by = request.user
        group.approved_at = timezone.now()
        group.save()
        
        serializer = AdminProjectGroupSerializer(group)
        return Response({
            'message': 'Group rejected',
            'data': serializer.data
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def details(self, request, pk=None):
        """GET /api/projects/admin/approval/{id}/details/"""
        self.check_admin_permission(request)
        
        try:
            group = ProjectGroup.objects.select_related(
                'supervisor', 'co_supervisor'
            ).prefetch_related('members__student').get(pk=pk)
        except ProjectGroup.DoesNotExist:
            return Response(
                {'error': 'Group not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = AdminProjectGroupSerializer(group)
        return Response(serializer.data)


# fyp-backend/projects/views.py - Replace the entire SupervisorGroupViewSet with this:

# =============================================================================
# ✅ SUPERVISOR DASHBOARD VIEWSET - FIXED (Using GenericViewSet)
# =============================================================================
class SupervisorGroupViewSet(viewsets.GenericViewSet):
    """
    Supervisor ke liye assigned groups fetch karne ka endpoint.
    Endpoint: GET /api/projects/groups/supervisor/
    """
    permission_classes = [IsAuthenticated]
    serializer_class = ProjectGroupSerializer  # Add this line

    def list(self, request):
        """GET /api/projects/groups/supervisor/"""
        user = request.user
        
        if user.user_type != 'supervisor':
            return Response(
                {'error': 'Only supervisors can access this endpoint'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            faculty = Faculty.objects.get(user=user)
        except Faculty.DoesNotExist:
            return Response(
                {'error': 'Faculty profile not found for this user'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        groups = ProjectGroup.objects.filter(
            models.Q(supervisor=faculty) | models.Q(co_supervisor=faculty),
            status__in=['idea_pitch', 'proposal_pending', 'proposal_approved', 
                    'in_progress', 'completed']
        ).prefetch_related(
            'members__student',  # ✅ Yeh important hai
            'supervisor',
            'co_supervisor'
        ).order_by('-created_at')
        
        serializer = self.get_serializer(groups, many=True, context={'request': request})
        
        # ✅ Debug print
        print(f"✅ Serialized groups: {len(serializer.data)} groups")
        for group_data in serializer.data:
            print(f"  Group {group_data.get('group_number')}: {len(group_data.get('members', []))} members")
        
        return Response({
            'count': groups.count(),
            'results': serializer.data  # ✅ Yeh sahi hai
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['get'])
    def members(self, request, pk=None):
        """GET /api/projects/groups/supervisor/{id}/members/"""
        try:
            group = ProjectGroup.objects.prefetch_related('members__student').get(pk=pk)
        except ProjectGroup.DoesNotExist:
            return Response({'error': 'Group not found'}, status=status.HTTP_404_NOT_FOUND)
        
        # Verify supervisor access
        faculty = Faculty.objects.filter(user=request.user).first()
        if not faculty or (group.supervisor != faculty and group.co_supervisor != faculty):
            return Response({'error': 'Access denied'}, status=status.HTTP_403_FORBIDDEN)
        
        members_data = []
        for member in group.members.all():
            members_data.append({
                'id': member.id,
                'full_name': member.full_name or member.student.get_full_name(),
                'email': member.student.email,
                'student_id': member.student.student_id,
                'role': member.role,
                'cgpa': float(member.cgpa),
                'credit_hours': member.earned_credit_hours,
                'contribution': member.contribution_percentage
            })
        
        return Response({'members': members_data})
    


# =============================================================================
# MEETING MINUTES VIEWSET
# =============================================================================

class MeetingMinuteViewSet(viewsets.ModelViewSet):
    """
    CRUD operations for meeting minutes.
    - Supervisor can create/update meetings for their groups
    - View all meetings for a group
    """
    serializer_class = MeetingMinuteSerializer
    permission_classes = [IsAuthenticated]
    
    def get_queryset(self):
        """
        Filter meetings based on user role:
        - Supervisors: Only their conducted meetings
        - Students: Only their group's meetings
        """
        user = self.request.user
        group_id = self.request.query_params.get('group_id')
        
        queryset = MeetingMinute.objects.all()
        
        if group_id:
            queryset = queryset.filter(group_id=group_id)
        
        if user.user_type == 'supervisor':
            try:
                faculty = user.faculty_profile
                queryset = queryset.filter(
                    models.Q(supervisor=faculty) | 
                    models.Q(group__co_supervisor=faculty)
                )
            except Faculty.DoesNotExist:
                return MeetingMinute.objects.none()
        
        elif user.user_type == 'student':
            # Students can only see their group's meetings
            queryset = queryset.filter(
                group__members__student=user
            ).distinct()
        
        return queryset.select_related('supervisor', 'group').prefetch_related('attendance_records')
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return MeetingMinuteCreateUpdateSerializer
        return MeetingMinuteSerializer
    
    def perform_create(self, serializer):
        group_id = self.request.data.get('group') or self.request.query_params.get('group_id')
        if group_id:
            try:
                group = ProjectGroup.objects.get(pk=group_id)
                # Supervisor auto-set karein
                if hasattr(self.request.user, 'faculty_profile'):
                    serializer.save(group=group, supervisor=self.request.user.faculty_profile)
                else:
                    serializer.save(group=group)
            except ProjectGroup.DoesNotExist:
                serializer.save()
        else:
            serializer.save()
    
    @action(detail=True, methods=['get'])
    def attendance_summary(self, request, pk=None):
        """
        Get attendance summary for a specific meeting.
        """
        meeting = self.get_object()
        attendance = meeting.attendance_records.all()
        
        summary = {
            'meeting_number': meeting.meeting_number,
            'date': meeting.date,
            'total_students': attendance.count(),
            'present': attendance.filter(status='present').count(),
            'absent': attendance.filter(status='absent').count(),
            'records': AttendanceLogSerializer(attendance, many=True).data
        }
        
        return Response(summary)


# =============================================================================
# ATTENDANCE SHEET VIEWSET (FP-5 Format)
# =============================================================================

class AttendanceSheetViewSet(viewsets.ViewSet):
    """
    Generate FP-5 style attendance sheet for a group.
    Shows all 16 meetings in grid format.
    """
    permission_classes = [IsAuthenticated]
    
    def retrieve(self, request, pk=None):
        """
        GET /api/projects/groups/{pk}/attendance-sheet/
        """
        try:
            group = ProjectGroup.objects.prefetch_related(
                'members__student',
                'meeting_minutes__attendance_records'
            ).get(pk=pk)
        except ProjectGroup.DoesNotExist:
            return Response(
                {'error': 'Group not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Verify access
        user = request.user
        if user.user_type == 'supervisor':
            try:
                faculty = user.faculty_profile
                if group.supervisor != faculty and group.co_supervisor != faculty:
                    return Response(
                        {'error': 'Access denied'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            except Faculty.DoesNotExist:
                return Response(
                    {'error': 'Faculty profile not found'},
                    status=status.HTTP_404_NOT_FOUND
                )
        
        # Build attendance sheet data
        members = group.members.all()
        meetings = group.meeting_minutes.all().order_by('meeting_number')
        
        # Create meeting number to data map
        meeting_map = {}
        for meeting in meetings:
            meeting_map[meeting.meeting_number] = {
                'date': meeting.date.strftime('%Y-%m-%d'),
                'attendance': {
                    rec.student.id: rec.status 
                    for rec in meeting.attendance_records.all()
                }
            }
        
        # Build member-wise attendance
        members_data = []
        for member in members:
            # ✅ Safely get full name
            student = member.student
            full_name = f"{student.first_name or ''} {student.last_name or ''}".strip() or student.email
            
            member_attendance = []
            total_present = 0
            
            for meeting_num in range(1, 17):
                if meeting_num in meeting_map:
                    att_status = meeting_map[meeting_num]['attendance'].get(
                        member.student.id, 'absent'
                    )
                    if att_status == 'present':
                        total_present += 1
                    member_attendance.append(att_status)
                else:
                    member_attendance.append(None)
            
            members_data.append({
                'student_id': member.student.id,
                'full_name': full_name,  # ✅ Fixed
                'odoo_id': member.student.student_id or 'N/A',
                'attendance': member_attendance,
                'total_present': total_present,
                'total_meetings': len(meetings)
            })
        
        # Build meetings summary
        meetings_summary = []
        for meeting_num in range(1, 17):
            if meeting_num in meeting_map:
                meetings_summary.append({
                    'meeting_number': meeting_num,
                    'date': meeting_map[meeting_num]['date'],
                    'conducted': True
                })
            else:
                meetings_summary.append({
                    'meeting_number': meeting_num,
                    'date': None,
                    'conducted': False
                })
        
        data = {
            'group_id': group.group_id,
            'group_number': group.group_number,
            'project_title': group.project_title,
            'supervisor_name': group.supervisor.full_name if group.supervisor else 'N/A',
            'semester': group.semester,
            'fydp_phase': group.fydp_phase,
            'members': members_data,
            'meetings_summary': meetings_summary,
            'total_meetings_conducted': len(meetings)
        }
        
        return Response(data)
    
    @action(detail=True, methods=['get'], url_path='export-excel')
    def export_excel(self, request, pk=None):
        """
        Export attendance sheet as Excel file.
        Requires: pip install openpyxl
        """
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.http import HttpResponse
        except ImportError:
            return Response(
                {'error': 'Excel export requires openpyxl. Install: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Get attendance sheet data
        response = self.retrieve(request, pk)
        if response.status_code != 200:
            return response
        
        data = response.data
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Attendance-{data['group_number']}"
        
        # Header styling
        header_fill = PatternFill(start_color="1e3a8a", end_color="1e3a8a", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"Attendance Sheet (FP-5) - {data['group_number']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Group info
        ws['A3'] = f"Project: {data['project_title']}"
        ws['A4'] = f"Supervisor: {data['supervisor_name']}"
        ws['A5'] = f"Semester: {data['semester']} | Phase: {data['fydp_phase']}"
        
        # Attendance table header
        ws.append([])  # Empty row
        headers = ['Seat No.', 'Student Name', 'Odoo ID'] + \
                  [f"Meeting {i}" for i in range(1, 17)] + \
                  ['Total']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Student data
        for idx, member in enumerate(data['members'], 8):
            ws.cell(row=idx, column=1, value=idx-7)  # Seat number
            ws.cell(row=idx, column=2, value=member['full_name'])
            ws.cell(row=idx, column=3, value=member['odoo_id'])
            
            for meeting_num, att_status in enumerate(member['attendance'], 4):
                if att_status == 'present':
                    ws.cell(row=idx, column=meeting_num, value='P')
                elif att_status == 'absent':
                    ws.cell(row=idx, column=meeting_num, value='A')
                else:
                    ws.cell(row=idx, column=meeting_num, value='-')
            
            # Total
            total_cell = ws.cell(row=idx, column=21, value=member['total_present'])
            total_cell.font = Font(bold=True)
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Attendance_{data["group_number"]}.xlsx'
        wb.save(response)
        
        return response
    


# =============================================================================
# ✅ STUDENT MEETING DATA VIEWSET
# =============================================================================

class StudentMeetingDataView(viewsets.ViewSet):
    """
    Student ke liye apni Group ki Attendance aur Tasks fetch karne ka endpoint.
    """
    permission_classes = [IsAuthenticated]
    
    def list(self, request):
        user = request.user
        
        try:
            # ✅ FIX: .get() ki jagah .filter() use karein aur pehla active group lein
            member = GroupMember.objects.filter(
                student=user,
                group__status__in=['idea_pitch', 'approved', 'proposal_pending', 'proposal_approved', 'in_progress', 'completed']
            ).select_related('group').first()
            
            if not member:
                return Response({
                    'error': 'No active group found',
                    'current_task': None,
                    'attendance': []
                }, status=status.HTTP_404_NOT_FOUND)
            
            group = member.group
            
            # 2. Meetings fetch karein
            meetings = MeetingMinute.objects.filter(group=group).order_by('meeting_number')
            
            # 3. Latest Task dhundein (Aakhri meeting se)
            current_task = None
            if meetings.exists():
                last_meeting = meetings.last()
                current_task = last_meeting.new_task
            
            # 4. Student ki Attendance Records fetch karein
            attendance_logs = AttendanceLog.objects.filter(
                meeting__group=group,
                student=user
            ).order_by('meeting__meeting_number')
            
            # Data structure build karein
            attendance_data = []
            for log in attendance_logs:
                attendance_data.append({
                    'meeting_number': log.meeting.meeting_number,
                    'date': log.meeting.date,
                    'status': log.status.capitalize(),  # ✅ Capitalize: 'present' -> 'Present'
                    'agenda': log.meeting.agenda,
                    'task_assigned': log.meeting.new_task 
                })
            
            response_data = {
                'group_number': group.group_number,
                'current_task': current_task,
                'attendance': attendance_data
            }
            
            return Response(response_data)
            
        except Exception as e:
            return Response({
                'error': str(e),
                'current_task': None,
                'attendance': []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





# =============================================================================
# ANNOUNCEMENT VIEWSET
# =============================================================================
class AnnouncementViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing announcements.
    
    - Public can view active announcements (READ only)
    - Admin/Staff can create, update, delete (FULL CRUD)
    """
    queryset = Announcement.objects.all()
    filter_backends = [filters.OrderingFilter]
    ordering_fields = ['priority', 'created_at', 'start_date']
    ordering = ['-priority', '-created_at']
    
    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return AnnouncementCreateUpdateSerializer
        return AnnouncementSerializer
    
    from .permissions import IsStudent, IsGroupMemberOrReadOnly, IsAdminUser
    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'active']:
            permission_classes = [permissions.AllowAny]
        else:
            permission_classes = [IsAdminUser]
        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Public users only see active announcements
        if not self.request.user.is_staff:
            queryset = queryset.filter(
                is_active=True,
                start_date__lte=timezone.now()
            )
            queryset = queryset.filter(
                models.Q(end_date__isnull=True) | 
                models.Q(end_date__gte=timezone.now())
            )
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        """
        GET /api/projects/announcements/active/
        Get all currently active announcements for homepage ticker
        """
        announcements = self.get_queryset().filter(is_active=True)
        serializer = self.get_serializer(announcements, many=True)
        return Response({
            'count': announcements.count(),
            'results': serializer.data
        })